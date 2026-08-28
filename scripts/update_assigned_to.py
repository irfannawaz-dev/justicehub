#!/usr/bin/env python3
"""
Reads column A (case_uid) and column AL (Responsible persons / assigned_to)
from all 5 hub sheets and outputs a JSON mapping {case_uid: assigned_to}.

Usage: python update_assigned_to.py <path_to_xlsx> > assigned_to_map.json
"""
import sys, json, io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else r'C:\xampp\htdocs\JusticeHub\Data Validation Sheet.xlsx'
SHEETS    = ['Hyderabad', 'Dadu', 'Islamabad', 'SBA', 'Sanghar']

def clean_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
mapping = {}

for sheet_name in SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] Sheet '{sheet_name}' not found, skipping.", file=sys.stderr)
        continue
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:        # skip header
        if not row[0]: continue
        case_uid    = clean_str(row[0])
        assigned_to = clean_str(row[37])  # column AL
        if case_uid and assigned_to:
            mapping[case_uid] = assigned_to

wb.close()
print(json.dumps(mapping, ensure_ascii=False))
